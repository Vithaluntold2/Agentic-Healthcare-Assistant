# database.py
# In-memory patient/doctor/appointment data layer.
# Patients are loaded from records.xlsx on first access.

import datetime
import uuid
import openpyxl
from typing import Optional
from src.config import RECORDS_PATH


# --- patients ---

_patients: dict[str, dict] = {}


def _seed_patients():
    """Load patients from the excel file into memory (called once)."""
    wb = openpyxl.load_workbook(RECORDS_PATH)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        phone = str(record.get("Phone_number", "")).strip()
        if not phone:
            continue
        pid = phone  # phone number as primary key
        if pid not in _patients:
            _patients[pid] = {
                "patient_id": pid,
                "name": record.get("Name", "Unknown"),
                "age": record.get("Age"),
                "gender": record.get("Gender"),
                "phone": phone,
                "email": record.get("Email"),
                "address": record.get("Address"),
                "summary": record.get("Summary", ""),
                "history": [],
            }


def get_all_patients() -> list[dict]:
    if not _patients:
        _seed_patients()
    return list(_patients.values())


def find_patient(query: str) -> Optional[dict]:
    """Search by name or phone. Case-insensitive partial match."""
    if not _patients:
        _seed_patients()
    query_lower = query.lower().strip()
    for pat in _patients.values():
        if query_lower in pat["name"].lower() or query_lower in pat["phone"]:
            return pat
    return None


def add_patient(name: str, age: int, gender: str, phone: str,
                address: str = "", email: str = "", summary: str = "") -> dict:
    """Register or update a patient. Phone is used as unique id."""
    if not _patients:
        _seed_patients()
    pid = phone.strip()
    _patients[pid] = {
        "patient_id": pid,
        "name": name,
        "age": age,
        "gender": gender,
        "phone": phone,
        "email": email,
        "address": address,
        "summary": summary,
        "history": _patients.get(pid, {}).get("history", []),
    }
    return _patients[pid]


def update_patient_summary(patient_id: str, new_summary: str) -> Optional[dict]:
    """Append notes to a patient's history."""
    if not _patients:
        _seed_patients()
    pat = _patients.get(patient_id)
    if pat is None:
        return None
    timestamp = datetime.datetime.now().isoformat()
    pat["history"].append({"date": timestamp, "note": new_summary})
    pat["summary"] = (pat["summary"] or "") + f"\n[{timestamp}] {new_summary}"
    return pat


# --- doctors ---

_doctors: list[dict] = [
    {
        "doctor_id": "D001",
        "name": "Dr. Priya Sharma",
        "specialty": "Nephrology",
        "available_slots": [
            {"date": "2026-03-16", "time": "09:00", "booked": False},
            {"date": "2026-03-16", "time": "11:00", "booked": False},
            {"date": "2026-03-17", "time": "10:00", "booked": False},
            {"date": "2026-03-18", "time": "14:00", "booked": False},
        ],
    },
    {
        "doctor_id": "D002",
        "name": "Dr. Anand Rao",
        "specialty": "General Medicine",
        "available_slots": [
            {"date": "2026-03-16", "time": "08:00", "booked": False},
            {"date": "2026-03-16", "time": "10:30", "booked": False},
            {"date": "2026-03-17", "time": "09:00", "booked": False},
            {"date": "2026-03-17", "time": "15:00", "booked": False},
        ],
    },
    {
        "doctor_id": "D003",
        "name": "Dr. Meena Iyer",
        "specialty": "Endocrinology",
        "available_slots": [
            {"date": "2026-03-16", "time": "10:00", "booked": False},
            {"date": "2026-03-17", "time": "11:00", "booked": False},
            {"date": "2026-03-18", "time": "09:30", "booked": False},
        ],
    },
    {
        "doctor_id": "D004",
        "name": "Dr. Rajesh Gupta",
        "specialty": "Cardiology",
        "available_slots": [
            {"date": "2026-03-16", "time": "14:00", "booked": False},
            {"date": "2026-03-17", "time": "08:30", "booked": False},
            {"date": "2026-03-18", "time": "10:00", "booked": False},
        ],
    },
    {
        "doctor_id": "D005",
        "name": "Dr. Sunita Verma",
        "specialty": "Pulmonology",
        "available_slots": [
            {"date": "2026-03-16", "time": "09:30", "booked": False},
            {"date": "2026-03-17", "time": "13:00", "booked": False},
            {"date": "2026-03-18", "time": "11:00", "booked": False},
        ],
    },
]


def get_all_doctors() -> list[dict]:
    return _doctors


def find_doctors_by_specialty(specialty: str) -> list[dict]:
    specialty_lower = specialty.lower().strip()
    return [d for d in _doctors if specialty_lower in d["specialty"].lower()]


def get_available_slots(doctor_id: str) -> list[dict]:
    """Unbooked slots for a given doctor."""
    for doc in _doctors:
        if doc["doctor_id"] == doctor_id:
            return [s for s in doc["available_slots"] if not s["booked"]]
    return []


# --- appointments ---

_appointments: list[dict] = []


def book_appointment(patient_id: str, doctor_id: str,
                     date: str, time: str) -> dict:
    """Try to book a slot. Returns {success: True/False, ...}."""
    for doc in _doctors:
        if doc["doctor_id"] == doctor_id:
            for slot in doc["available_slots"]:
                if slot["date"] == date and slot["time"] == time:
                    if slot["booked"]:
                        return {"success": False, "error": "Slot already booked"}
                    slot["booked"] = True
                    appt = {
                        "appointment_id": str(uuid.uuid4())[:8],
                        "patient_id": patient_id,
                        "doctor_id": doctor_id,
                        "doctor_name": doc["name"],
                        "specialty": doc["specialty"],
                        "date": date,
                        "time": time,
                        "status": "confirmed",
                        "created_at": datetime.datetime.now().isoformat(),
                    }
                    _appointments.append(appt)
                    return {"success": True, "appointment": appt}
            return {"success": False, "error": f"Slot {date} {time} not found"}
    return {"success": False, "error": f"Doctor {doctor_id} not found"}


def get_appointments(patient_id: Optional[str] = None) -> list[dict]:
    if patient_id:
        return [a for a in _appointments if a["patient_id"] == patient_id]
    return _appointments


def cancel_appointment(appointment_id: str) -> dict:
    """Cancel appointment and free the slot back up."""
    for i, appt in enumerate(_appointments):
        if appt["appointment_id"] == appointment_id:
            for doc in _doctors:
                if doc["doctor_id"] == appt["doctor_id"]:
                    for slot in doc["available_slots"]:
                        if slot["date"] == appt["date"] and slot["time"] == appt["time"]:
                            slot["booked"] = False
            appt["status"] = "cancelled"
            return {"success": True, "message": f"Appointment {appointment_id} cancelled"}
    return {"success": False, "error": "Appointment not found"}
